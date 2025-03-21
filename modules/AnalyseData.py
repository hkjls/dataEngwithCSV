from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import filedialog

class Analyser:
    def __init__(self):
        self._newAnalyser = None
        
        root = tk.Tk()
        root.withdraw()
        self._file_path = filedialog.askopenfilename(
            title="Select file",
            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*"))
        )
        
        if not self._file_path:
            raise ValueError("No file selected")
        
        if type(self._file_path) == str:
            self.wb = load_workbook(self._file_path)
            # self.ws = self.wb.active
            
    def copySheets(self):
        self._newAnalyser = Workbook()
        self._newAnalyser.remove(self._newAnalyser.active)
        sheets = filter(lambda x:x != "Collage saturne" ,self.wb.sheetnames)
        
        for sheet in sheets:
            ws = self._newAnalyser.create_sheet(sheet)
            ws = self._newAnalyser[sheet]
            for row in self.wb[sheet].iter_rows():
                for cell in row:
                    ws[cell.coordinate].value = cell.value
                    
        self._newAnalyser.create_sheet(title="Collage saturne", index=2)
        
    def saveAnalyser(self)-> str:
        file = filedialog.asksaveasfilename(
            title="Save file",
            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*"))
        )
        
        if not file:
            raise ValueError("No file selected")
        
        self._newAnalyser.save(f"{file}.xlsx")
        
        return f"{file}.xlsx"
        
    
    def copyContent(self, data: classmethod, sheet_name):
        sheet = self._newAnalyser[sheet_name]
        for r_idx, row in enumerate(data.itertuples(index=False, name=None), 1):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, str) and value.replace(",", "").replace(".", "").isdigit():
                        if "," in value:
                            value = value.replace(",", ".")  # Convert comma decimal separator to dot
                        value = float(value) if "." in value else int(value)  # Convert to number
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                cell.number_format = "General"
        return

        
    