import openpyxl
import tkinter as tk
from tkinter import filedialog
import pandas as pd

class handleContent:
    def __init__(self) -> None:
        
        root = tk.Tk()
        root.withdraw()
        self._wb = None
        self._ws = None
    
    def set_filename(self, filetype: str):
        filename = filedialog.askopenfilename(
                title="Select the file",
                filetypes=(("files", filetype), ("all files", "*.*"))
        )
            
        return filename
        
    def Delete(self,filename:str, sheetName: str, start_rows: int = 1, start_column: int = 1) -> None:        
        self._wb = openpyxl.load_workbook(filename)
        self._ws = self._wb[sheetName]
        
        self._ws.delete_rows(start_rows, self._ws.max_row)
        self._ws.delete_cols(start_column, self._ws.max_column)
            
        self._wb.save(filename)
        
    def Insert(self, dt: pd.core.frame.DataFrame, targetfile: str, sheet_name: str):
        with pd.ExcelWriter(targetfile, mode='a', if_sheet_exists='overlay') as writer:
            dt.to_excel(writer, sheet_name=sheet_name, startrow=0, index=False, header=False)