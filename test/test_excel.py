import os
import pandas as pd
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import filedialog

def csv_to_excel_and_process(csv_file: str, excel_filepath: str) -> pd.DataFrame:
    """
    Reads CSV data (with semicolon separator), writes it to a new Excel file (appending '_result' to filename),
    while preserving all sheets from the original file, including formulas. 
    If "Collage saturne" exists, it is cleared before writing new data (without headers).
    Returns a DataFrame from the "évolution débit" sheet.
    """
    try:
        # Read the CSV content into a DataFrame
        df_csv = pd.read_csv(csv_file, sep=";", names=range(6))

        # Create a new file path with "_result" appended before the extension
        # base, ext = os.path.splitext(excel_filepath)
        # new_excel_filepath = f"{base}_result{ext}"

        # Load the original workbook
        try:
            workbook = load_workbook(excel_filepath)
        except FileNotFoundError:
            print(f"File not found: {excel_filepath}. Creating a new one.")
            workbook = Workbook()  # Create a new workbook

        # Create a new workbook to save the result
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)  # Remove default empty sheet

        # Copy all sheets from the original workbook, preserving formulas
        for sheet_name in workbook.sheetnames:
            source_sheet = workbook[sheet_name]

            # If it's "Collage saturne", clear its content before copying
            if sheet_name == "Collage saturne":
                new_sheet = new_workbook.create_sheet(title=sheet_name)
            else:
                new_sheet = new_workbook.create_sheet(title=sheet_name)

                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_cell = new_sheet[cell.coordinate]
                        new_cell.value = cell.value  # Preserve formulas, values, and formatting

        # Write new data to "Collage saturne" (without headers)
        if "Collage saturne" in new_workbook.sheetnames:
            sheet = new_workbook["Collage saturne"]
            for r_idx, row in enumerate(df_csv.itertuples(index=False, name=None), start=1):
                for c_idx, value in enumerate(row, start=1):
                    # Convert numbers properly
                    if isinstance(value, str) and value.replace(",", "").replace(".", "").isdigit():
                        if "," in value:
                            value = value.replace(",", ".")  # Convert comma decimal separator to dot
                        value = float(value) if "." in value else int(value)  # Convert to number

                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    cell.number_format = "General"  # Ensure standard format


        # Save the new workbook
        new_excel_filepath = filedialog.asksaveasfilename(
            title="Save file",
            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*"))
        )
        new_workbook.save(new_excel_filepath)

        # Read the "évolution débit" tab into a DataFrame from the new file
        try:
            df_result = pd.read_excel(new_excel_filepath, sheet_name="évolution débit", engine="openpyxl")
            return df_result
        except ValueError:
            print(f"Sheet 'évolution débit' not found in {new_excel_filepath}")
            return None
    except pd.errors.ParserError as e:
        print(f"Error parsing CSV: {e}")
        return None
    except FileNotFoundError:
        print(f"File not found: {csv_file}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None


csv_file = filedialog.askopenfilename(title="Select CSV file Source", filetypes=(("csv files", "*.csv"), ("all files", "*.*")))
excel_file = filedialog.askopenfilename(title="Select Excel file Source", filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))   

result_df = csv_to_excel_and_process(csv_file, excel_file)

if result_df is not None:
    print("DataFrame from 'évolution débit' tab:")
    print(result_df)
else:
    print("An error occurred while processing the files.")